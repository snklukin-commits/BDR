"""
AI mail assistant for a BSS/IMAP mailbox.

Runs as a small Flask app:
- checks IMAP automatically;
- stores only metadata, body text and extracted attachment text in SQLite;
- summarizes new mail;
- generates recommended replies;
- can send a reply through SMTP only after an explicit API/UI action.

Secrets must be provided through environment variables. Do not commit passwords.
"""

from __future__ import annotations

import email
import html
import imaplib
import io
import json
import os
import re
import smtplib
import sqlite3
import ssl
import threading
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from email.header import decode_header, make_header
from email.message import EmailMessage
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from flask import Flask, jsonify, request, render_template_string

try:
    from openai import OpenAI
except Exception:  # pragma: no cover
    OpenAI = None  # type: ignore

try:
    from pdfminer.high_level import extract_text as pdf_extract_text
except Exception:  # pragma: no cover
    pdf_extract_text = None

try:
    from docx import Document
except Exception:  # pragma: no cover
    Document = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = os.getenv("DB_PATH", str(BASE_DIR / "mail_ai.sqlite3"))

IMAP_HOST = os.getenv("IMAP_HOST", "mx03.bsspharm.ru")
IMAP_PORT = int(os.getenv("IMAP_PORT", "993"))
IMAP_USE_SSL = os.getenv("IMAP_USE_SSL", "true").lower() == "true"

SMTP_HOST = os.getenv("SMTP_HOST", "mx03.bsspharm.ru")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_STARTTLS = os.getenv("SMTP_STARTTLS", "true").lower() == "true"

MAIL_USER = os.getenv("MAIL_USER") or os.getenv("EMAIL_USER")
MAIL_PASSWORD = os.getenv("MAIL_PASSWORD") or os.getenv("EMAIL_PASSWORD")
MAIL_FROM = os.getenv("MAIL_FROM") or MAIL_USER

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
CHECK_INTERVAL_SECONDS = int(os.getenv("CHECK_INTERVAL_SECONDS", "300"))
MAX_EMAILS_PER_CHECK = int(os.getenv("MAX_EMAILS_PER_CHECK", "30"))
AUTO_EMAIL_REPORT_TO = os.getenv("AUTO_EMAIL_REPORT_TO", "").strip()
AUTO_EMAIL_REPORT_ENABLED = os.getenv("AUTO_EMAIL_REPORT_ENABLED", "false").lower() == "true"

app = Flask(__name__)

INDEX_HTML = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>AI-почта</title>
  <style>
    :root { color-scheme: light dark; }
    body { margin: 0; font-family: Arial, sans-serif; background: #f5f5f5; color: #111; }
    header { padding: 16px 20px; background: #111827; color: white; }
    main { display: grid; grid-template-columns: 380px 1fr; gap: 0; min-height: calc(100vh - 60px); }
    aside { border-right: 1px solid #ddd; background: white; overflow: auto; }
    section { padding: 20px; overflow: auto; }
    button { border: 0; background: #1f2937; color: white; padding: 9px 12px; border-radius: 8px; cursor: pointer; }
    button.secondary { background: #e5e7eb; color: #111; }
    input, textarea { width: 100%; box-sizing: border-box; padding: 10px; border: 1px solid #ccc; border-radius: 8px; }
    textarea { min-height: 180px; }
    .toolbar { display: flex; gap: 8px; padding: 12px; border-bottom: 1px solid #ddd; }
    .mail { padding: 12px 14px; border-bottom: 1px solid #eee; cursor: pointer; }
    .mail:hover { background: #f3f4f6; }
    .mail.active { background: #e5e7eb; }
    .subject { font-weight: 700; margin-bottom: 4px; }
    .meta { color: #6b7280; font-size: 12px; }
    .badge { display: inline-block; padding: 2px 6px; border-radius: 999px; background: #e5e7eb; font-size: 12px; margin-right: 6px; }
    .card { background: white; border: 1px solid #ddd; border-radius: 12px; padding: 16px; margin-bottom: 16px; }
    pre { white-space: pre-wrap; word-wrap: break-word; }
    @media (prefers-color-scheme: dark) {
      body { background: #0b0f17; color: #f9fafb; }
      aside, .card { background: #111827; border-color: #374151; }
      .mail { border-color: #1f2937; }
      .mail:hover, .mail.active { background: #1f2937; }
      input, textarea { background: #0b0f17; color: #f9fafb; border-color: #374151; }
      button.secondary { background: #374151; color: #f9fafb; }
    }
  </style>
</head>
<body>
<header><strong>AI-почта</strong> · IMAP/SMTP · mx03.bsspharm.ru</header>
<main>
  <aside>
    <div class="toolbar">
      <button onclick="syncNow()">Проверить</button>
      <button class="secondary" onclick="loadReport()">Сводка</button>
    </div>
    <div id="mailList"></div>
  </aside>
  <section>
    <div id="content" class="card">Выберите письмо слева или нажмите «Проверить».</div>
  </section>
</main>
<script>
let selectedId = null;
async function api(url, options) {
  const res = await fetch(url, options);
  if (!res.ok) throw new Error(await res.text());
  return await res.json();
}
async function loadMails() {
  const data = await api('/api/emails');
  const list = document.getElementById('mailList');
  list.innerHTML = '';
  data.emails.forEach(m => {
    const el = document.createElement('div');
    el.className = 'mail' + (m.id === selectedId ? ' active' : '');
    el.innerHTML = `<div class="subject">${escapeHtml(m.subject || '(без темы)')}</div>
      <div class="meta">${escapeHtml(m.sender || '')}</div>
      <div class="meta">${escapeHtml(m.date || '')} · ${m.has_attachments ? 'вложения' : 'без вложений'}</div>`;
    el.onclick = () => openMail(m.id);
    list.appendChild(el);
  });
}
async function syncNow() {
  document.getElementById('content').innerHTML = '<div class="card">Проверяю почту...</div>';
  const data = await api('/api/sync', {method: 'POST'});
  await loadMails();
  document.getElementById('content').innerHTML = `<div class="card">Новых писем: ${data.new_count}</div>`;
}
async function openMail(id) {
  selectedId = id;
  await loadMails();
  const m = await api('/api/email/' + id);
  document.getElementById('content').innerHTML = `
    <div class="card">
      <h2>${escapeHtml(m.subject || '(без темы)')}</h2>
      <p><strong>От:</strong> ${escapeHtml(m.sender || '')}</p>
      <p><strong>Дата:</strong> ${escapeHtml(m.date || '')}</p>
      <p>${m.has_attachments ? '<span class="badge">есть вложения</span>' : ''}</p>
      <h3>Тело письма</h3>
      <pre>${escapeHtml(m.body || '')}</pre>
      <h3>Текст из вложений</h3>
      <pre>${escapeHtml(m.attachments_text || 'Нет извлечённого текста')}</pre>
      <p><button onclick="generateReply(${id})">Сгенерировать ответ</button> <button class="secondary" onclick="summarizeEmail(${id})">Сводка письма</button></p>
    </div>
    <div id="aiBox" class="card"></div>`;
}
async function generateReply(id) {
  document.getElementById('aiBox').innerHTML = 'Генерирую ответ...';
  const data = await api('/api/email/' + id + '/reply', {method: 'POST'});
  document.getElementById('aiBox').innerHTML = `<h3>Рекомендованный ответ</h3><textarea id="replyText">${escapeHtml(data.reply)}</textarea><p><input id="replyTo" value="${escapeHtml(data.reply_to || '')}" placeholder="Кому" /></p><p><button onclick="sendReply(${id})">Отправить через SMTP</button></p>`;
}
async function summarizeEmail(id) {
  document.getElementById('aiBox').innerHTML = 'Делаю сводку...';
  const data = await api('/api/email/' + id + '/summary', {method: 'POST'});
  document.getElementById('aiBox').innerHTML = `<h3>Сводка</h3><pre>${escapeHtml(data.summary)}</pre>`;
}
async function sendReply(id) {
  const body = document.getElementById('replyText').value;
  const to = document.getElementById('replyTo').value;
  const data = await api('/api/email/' + id + '/send-reply', {method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({to, body})});
  document.getElementById('aiBox').innerHTML += `<p><strong>${escapeHtml(data.status)}</strong></p>`;
}
async function loadReport() {
  document.getElementById('content').innerHTML = '<div class="card">Готовлю сводку...</div>';
  const data = await api('/api/report', {method: 'POST'});
  document.getElementById('content').innerHTML = `<div class="card"><h2>Сводка новых писем</h2><pre>${escapeHtml(data.report)}</pre></div>`;
}
function escapeHtml(s) {
  return String(s || '').replace(/[&<>'"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[c]));
}
loadMails();
</script>
</body>
</html>
"""


@dataclass
class ParsedMail:
    uid: str
    subject: str
    sender: str
    reply_to: str
    date: str
    body: str
    attachments_text: str
    attachments_meta: list[dict[str, Any]]


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS emails (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uid TEXT UNIQUE NOT NULL,
                subject TEXT,
                sender TEXT,
                reply_to TEXT,
                date TEXT,
                body TEXT,
                attachments_text TEXT,
                attachments_meta TEXT,
                has_attachments INTEGER DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )


def decode_mime(value: str | None) -> str:
    if not value:
        return ""
    try:
        return str(make_header(decode_header(value)))
    except Exception:
        return value


def strip_html(value: str) -> str:
    value = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", value)
    value = re.sub(r"(?is)<br\s*/?>", "\n", value)
    value = re.sub(r"(?is)</p>", "\n", value)
    value = re.sub(r"(?is)<.*?>", " ", value)
    value = html.unescape(value)
    return re.sub(r"[ \t]+", " ", value).strip()


def payload_to_text(part: email.message.Message) -> str:
    payload = part.get_payload(decode=True) or b""
    charset = part.get_content_charset() or "utf-8"
    try:
        return payload.decode(charset, errors="ignore")
    except LookupError:
        return payload.decode("utf-8", errors="ignore")


def extract_body(msg: email.message.Message) -> str:
    plain_parts: list[str] = []
    html_parts: list[str] = []
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_disposition() == "attachment":
                continue
            ctype = part.get_content_type()
            if ctype == "text/plain":
                plain_parts.append(payload_to_text(part))
            elif ctype == "text/html":
                html_parts.append(strip_html(payload_to_text(part)))
    else:
        ctype = msg.get_content_type()
        if ctype == "text/html":
            html_parts.append(strip_html(payload_to_text(msg)))
        else:
            plain_parts.append(payload_to_text(msg))
    body = "\n\n".join(p.strip() for p in plain_parts if p.strip())
    if not body:
        body = "\n\n".join(p.strip() for p in html_parts if p.strip())
    return body.strip()


def extract_attachment_text(filename: str, data: bytes) -> str:
    suffix = Path(filename or "attachment").suffix.lower()
    try:
        if suffix == ".pdf" and pdf_extract_text:
            return pdf_extract_text(io.BytesIO(data))[:20000]
        if suffix == ".docx" and Document:
            document = Document(io.BytesIO(data))
            return "\n".join(p.text for p in document.paragraphs)[:20000]
        if suffix in {".xlsx", ".xlsm"} and load_workbook:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines: list[str] = []
            for ws in wb.worksheets[:5]:
                lines.append(f"Лист: {ws.title}")
                for row in ws.iter_rows(max_row=100, values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append("\t".join(cells))
            return "\n".join(lines)[:20000]
        if suffix in {".txt", ".csv"}:
            return data.decode("utf-8", errors="ignore")[:20000]
    except Exception as exc:
        return f"Не удалось извлечь текст из {filename}: {exc}"
    return ""


def parse_message(uid: str, raw: bytes) -> ParsedMail:
    msg = email.message_from_bytes(raw)
    subject = decode_mime(msg.get("Subject"))
    sender = decode_mime(msg.get("From"))
    reply_to = decode_mime(msg.get("Reply-To")) or sender
    date_header = msg.get("Date")
    try:
        date = parsedate_to_datetime(date_header).astimezone().isoformat(timespec="minutes") if date_header else ""
    except Exception:
        date = date_header or ""
    body = extract_body(msg)

    attachment_blocks: list[str] = []
    attachments_meta: list[dict[str, Any]] = []
    for part in msg.walk():
        if part.get_content_disposition() != "attachment":
            continue
        filename = decode_mime(part.get_filename()) or "attachment"
        data = part.get_payload(decode=True) or b""
        attachments_meta.append({"filename": filename, "size": len(data), "content_type": part.get_content_type()})
        extracted = extract_attachment_text(filename, data)
        if extracted:
            attachment_blocks.append(f"=== {filename} ===\n{extracted}")

    return ParsedMail(
        uid=uid,
        subject=subject,
        sender=sender,
        reply_to=reply_to,
        date=date,
        body=body,
        attachments_text="\n\n".join(attachment_blocks),
        attachments_meta=attachments_meta,
    )


def connect_imap() -> imaplib.IMAP4:
    if not MAIL_USER or not MAIL_PASSWORD:
        raise RuntimeError("MAIL_USER/MAIL_PASSWORD не заданы")
    if IMAP_USE_SSL:
        client = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    else:
        client = imaplib.IMAP4(IMAP_HOST, IMAP_PORT)
    client.login(MAIL_USER, MAIL_PASSWORD)
    return client


def save_mail(mail: ParsedMail) -> bool:
    with db() as conn:
        try:
            conn.execute(
                """
                INSERT INTO emails(uid, subject, sender, reply_to, date, body, attachments_text, attachments_meta, has_attachments, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    mail.uid,
                    mail.subject,
                    mail.sender,
                    mail.reply_to,
                    mail.date,
                    mail.body,
                    mail.attachments_text,
                    json.dumps(mail.attachments_meta, ensure_ascii=False),
                    1 if mail.attachments_meta else 0,
                    datetime.now(timezone.utc).isoformat(timespec="seconds"),
                ),
            )
            return True
        except sqlite3.IntegrityError:
            return False


def sync_mailbox() -> int:
    client = connect_imap()
    try:
        status, _ = client.select("INBOX")
        if status != "OK":
            raise RuntimeError("Не удалось открыть INBOX")
        status, data = client.uid("SEARCH", None, "UNSEEN")
        if status != "OK":
            raise RuntimeError("Не удалось выполнить IMAP SEARCH")
        uids = data[0].split()[-MAX_EMAILS_PER_CHECK:]
        new_count = 0
        for raw_uid in uids:
            uid = raw_uid.decode()
            status, msg_data = client.uid("FETCH", raw_uid, "(RFC822)")
            if status != "OK" or not msg_data:
                continue
            raw = msg_data[0][1]
            if not isinstance(raw, bytes):
                continue
            parsed = parse_message(uid, raw)
            if save_mail(parsed):
                new_count += 1
        return new_count
    finally:
        try:
            client.close()
        except Exception:
            pass
        client.logout()


def openai_client() -> Any:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY не задан. Подписка ChatGPT не заменяет API-ключ.")
    if OpenAI is None:
        raise RuntimeError("Пакет openai не установлен")
    return OpenAI(api_key=OPENAI_API_KEY)


def ai_complete(system: str, user: str, max_tokens: int = 900) -> str:
    client = openai_client()
    result = client.chat.completions.create(
        model=OPENAI_MODEL,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user[:45000]}],
        temperature=0.2,
        max_tokens=max_tokens,
    )
    return (result.choices[0].message.content or "").strip()


def get_email_row(email_id: int) -> sqlite3.Row | None:
    with db() as conn:
        return conn.execute("SELECT * FROM emails WHERE id = ?", (email_id,)).fetchone()


def build_email_context(row: sqlite3.Row) -> str:
    return f"""
Тема: {row['subject']}
От: {row['sender']}
Дата: {row['date']}

Тело письма:
{row['body']}

Извлеченный текст из вложений:
{row['attachments_text'] or 'нет'}
""".strip()


def generate_reply(row: sqlite3.Row) -> str:
    return ai_complete(
        "Ты помощник для рабочей почты аптечной сети. Составь короткий, конкретный, деловой ответ на русском. Не выдумывай факты. Если нужно уточнение — прямо напиши, что нужно уточнить.",
        build_email_context(row),
        max_tokens=600,
    )


def summarize(row: sqlite3.Row) -> str:
    return ai_complete(
        "Сделай рабочую сводку письма на русском: суть, что требуется, дедлайн, риски, какие данные есть во вложениях, рекомендуемое действие.",
        build_email_context(row),
        max_tokens=700,
    )


def make_report(limit: int = 20) -> str:
    with db() as conn:
        rows = conn.execute("SELECT * FROM emails ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    if not rows:
        return "Новых сохранённых писем нет."
    joined = "\n\n---\n\n".join(build_email_context(row) for row in rows)
    return ai_complete(
        "Сделай сводный отчёт по новым рабочим письмам на русском. Разделы: срочно, требует ответа, просто к сведению, данные из вложений, рекомендуемые ответы тезисно. Не придумывай то, чего нет в письмах.",
        joined,
        max_tokens=1200,
    )


def send_mail(to_addr: str, subject: str, body: str, reply_to_message: sqlite3.Row | None = None) -> None:
    if not MAIL_USER or not MAIL_PASSWORD or not MAIL_FROM:
        raise RuntimeError("MAIL_USER/MAIL_PASSWORD/MAIL_FROM не заданы")
    msg = EmailMessage()
    msg["From"] = MAIL_FROM
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.set_content(body)
    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
        if SMTP_STARTTLS:
            server.starttls(context=context)
        server.login(MAIL_USER, MAIL_PASSWORD)
        server.send_message(msg)


def background_worker() -> None:
    while True:
        try:
            new_count = sync_mailbox()
            if AUTO_EMAIL_REPORT_ENABLED and AUTO_EMAIL_REPORT_TO and new_count:
                report = make_report()
                send_mail(AUTO_EMAIL_REPORT_TO, "AI-сводка новых рабочих писем", report)
                with db() as conn:
                    conn.execute("INSERT INTO reports(report, created_at) VALUES (?, ?)", (report, datetime.now(timezone.utc).isoformat(timespec="seconds")))
        except Exception as exc:
            print(f"background_worker error: {exc}")
        time.sleep(CHECK_INTERVAL_SECONDS)


@app.get("/")
def index() -> str:
    return render_template_string(INDEX_HTML)


@app.get("/api/health")
def health() -> Any:
    return jsonify({
        "status": "ok",
        "imap_host": IMAP_HOST,
        "imap_port": IMAP_PORT,
        "smtp_host": SMTP_HOST,
        "smtp_port": SMTP_PORT,
        "mail_user_set": bool(MAIL_USER),
        "openai_key_set": bool(OPENAI_API_KEY),
    })


@app.post("/api/sync")
def api_sync() -> Any:
    return jsonify({"new_count": sync_mailbox()})


@app.get("/api/emails")
def api_emails() -> Any:
    with db() as conn:
        rows = conn.execute(
            "SELECT id, subject, sender, date, has_attachments FROM emails ORDER BY id DESC LIMIT 100"
        ).fetchall()
    return jsonify({"emails": [dict(row) for row in rows]})


@app.get("/api/email/<int:email_id>")
def api_email(email_id: int) -> Any:
    row = get_email_row(email_id)
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(dict(row))


@app.post("/api/email/<int:email_id>/reply")
def api_reply(email_id: int) -> Any:
    row = get_email_row(email_id)
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify({"reply": generate_reply(row), "reply_to": row["reply_to"]})


@app.post("/api/email/<int:email_id>/summary")
def api_summary(email_id: int) -> Any:
    row = get_email_row(email_id)
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify({"summary": summarize(row)})


@app.post("/api/email/<int:email_id>/send-reply")
def api_send_reply(email_id: int) -> Any:
    row = get_email_row(email_id)
    if not row:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(silent=True) or {}
    to_addr = data.get("to") or row["reply_to"]
    body = data.get("body") or ""
    if not to_addr or not body.strip():
        return jsonify({"error": "to/body required"}), 400
    send_mail(to_addr, "Re: " + (row["subject"] or ""), body, row)
    return jsonify({"status": "Отправлено"})


@app.post("/api/report")
def api_report() -> Any:
    report = make_report()
    with db() as conn:
        conn.execute("INSERT INTO reports(report, created_at) VALUES (?, ?)", (report, datetime.now(timezone.utc).isoformat(timespec="seconds")))
    return jsonify({"report": report})


if __name__ == "__main__":
    init_db()
    if MAIL_USER and MAIL_PASSWORD:
        threading.Thread(target=background_worker, daemon=True).start()
    else:
        print("MAIL_USER/MAIL_PASSWORD не заданы. Автопроверка отключена.")
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")))
